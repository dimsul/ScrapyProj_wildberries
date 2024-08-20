from openpyxl import load_workbook
from dataclasses import dataclass

from xlsx_reader import XlsxReader
from params import XLSX_WORK_SHEET


@dataclass
class XlsxWriter(XlsxReader):
    """Запись файла в xlsx файл"""

    row_counter = 2

    value: [int, float]

    def get_list_of_requests(self):
        pass

    def set_data_to_xlsx_file(self):

        self.__insert_data(self.filename, self.value)

    @classmethod
    def __insert_data(cls, filename, value):
        """Запись данных в файл"""

        try:

            wb = load_workbook(f'{filename}')
            ws = wb[XLSX_WORK_SHEET]
            ws[f'Z{cls.row_counter}'] = value
            cls.row_counter += 1
            wb.save(f'{filename}')

        except Exception as err:

            with open(f"./error_log/write_to_xlsx_error.txt", 'a') as file:
                file.write(f'{err}')
