from openpyxl import load_workbook
from dataclasses import dataclass
from os import path

from params import XLSX_WORK_SHEET


@dataclass
class XlsxReader:
    """Ридер xlsx файлов для получения поисковых запросов"""

    filename: str

    def __post_init__(self):

        if not path.isfile(self.filename):
            raise FileNotFoundError

    def get_list_of_requests(self):
        """Получаем список запросов и возвращаем их в программу"""

        try:
            ws = load_workbook(f'{self.filename}')
            requests = []
            row = 2
            while True:
                if ws[XLSX_WORK_SHEET].cell(row, 25).value is None:
                    break
                else:
                    requests.append(ws[XLSX_WORK_SHEET].cell(row, 25).value)
                    row += 1
        except Exception:
            raise FileExistsError

        return requests
